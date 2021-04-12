# _*_ coding:utf-8 _*_
# File Name: operation_excel
# Author： Emily
# Date:  2021/4/9  16:12
# Description : 操作Excel表的基础方法


import openpyxl

# 创建一个工作簿
wb = openpyxl.Workbook()
# 创建一个test_case的sheet表单
wb.create_sheet('test_case')
# 保存为一个xlsx格式的文件
wb.save('cases.xlsx')
# 读取Excel中的数据
# 第一步：打开工作簿
wb = openpyxl.load_workbook('cases.xlsx')
# 第二步：选取表单
sh = wb['test_case']
# 第三部：读取数据
# 参数 row:行 column：列
# 读取第一行，第一列的数据
ce = sh.cell(row=1, column=1)
print(ce.value)
# 按行读取数据 list(sh.rows)
# 按行读取数据，去掉第一行的表头信息数据
print(list(sh.rows)[1:])
for cases in list(sh.rows)[1:]:
    case_id = cases[0].value
    case_excepted = cases[1].value
    case_data = cases[2].value
    print(case_excepted, case_data)
# 关闭工作薄
wb.close()
